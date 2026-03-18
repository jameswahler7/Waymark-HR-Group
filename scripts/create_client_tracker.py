"""
create_client_tracker.py
Waymark HR Group LLC

Creates the client tracker spreadsheet in data/spreadsheets/.
Run once to initialize. Re-run to regenerate from scratch.
"""

from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).resolve().parents[1] / "data" / "spreadsheets" / "Waymark_Client_Tracker.xlsx"

# Brand colors (hex, no #)
NAVY   = "1A3A5C"
TEAL   = "007A87"
LGRAY  = "F4F6F8"
WHITE  = "FFFFFF"
TEXT   = "2C2C2C"
MUTED  = "888888"
ACCENT = "E8F4F6"

def col_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def make_header_row(ws, row, headers, fill_color, font_color=WHITE, font_size=10):
    for col, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name="Calibri", bold=True, size=font_size, color=font_color)
        cell.fill = col_fill(fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = width


def build_clients_sheet(wb):
    ws = wb.active
    ws.title = "Client Tracker"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 36

    # Row 1: Title
    ws.merge_cells("A1:N1")
    title = ws["A1"]
    title.value = "WAYMARK HR GROUP LLC — CLIENT TRACKER"
    title.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title.fill = col_fill(NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Subtitle
    ws.merge_cells("A2:N2")
    sub = ws["A2"]
    sub.value = f"Last Updated: {date.today().strftime('%B %d, %Y')}   |   Waymark HR Group LLC   |   716-225-6347   |   Jamie.Wahler@waymarkhrgroup.com"
    sub.font = Font(name="Calibri", size=9, color=WHITE)
    sub.fill = col_fill(TEAL)
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3: Column headers
    headers = [
        ("Client ID",          10),
        ("Company Name",       22),
        ("Industry",           18),
        ("Size\n(Employees)",  12),
        ("State",              10),
        ("Primary Contact",    20),
        ("Contact Email",      26),
        ("Contact Phone",      16),
        ("Services",           28),
        ("Status",             14),
        ("Start Date",         13),
        ("Contract Value",     15),
        ("Documents",          22),
        ("Notes",              30),
    ]
    make_header_row(ws, 3, headers, NAVY)

    # Sample data rows
    sample_data = [
        ["WHG-001", "Waymark HR Group LLC", "HR Consulting", 10, "New York",
         "Jamie Wahler", "Jamie.Wahler@waymarkhrgroup.com", "716-225-6347",
         "Policy Creation, Onboarding", "Active", "2026-03-18", "$0",
         "WaymarkHRGroupLLC_HR_Policies_2026-03-18.docx", "Internal account"],

        ["WHG-002", "Acme Logistics", "Logistics", 50, "Ohio",
         "", "", "",
         "Policy Creation", "Active", "2026-03-18", "",
         "AcmeLogistics_HR_Policies_2026-03-18.docx", ""],

        ["WHG-003", "", "", "", "", "", "", "", "", "Prospect", "", "", "", ""],
        ["WHG-004", "", "", "", "", "", "", "", "", "Prospect", "", "", "", ""],
        ["WHG-005", "", "", "", "", "", "", "", "", "Prospect", "", "", "", ""],
    ]

    status_colors = {
        "Active":    "D6F0D6",
        "Prospect":  "FFF3CD",
        "On Hold":   "FFE0CC",
        "Completed": "E0E0E0",
    }

    for i, row_data in enumerate(sample_data, start=4):
        ws.row_dimensions[i].height = 18
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = Font(name="Calibri", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="center" if col in (1, 4, 5, 10, 11) else "left")
            cell.border = thin_border()
            # Alternate row shading
            bg = ACCENT if i % 2 == 0 else WHITE
            cell.fill = col_fill(bg)

        # Color-code the Status cell
        status_val = row_data[9]
        if status_val in status_colors:
            status_cell = ws.cell(row=i, column=10)
            status_cell.fill = col_fill(status_colors[status_val])
            status_cell.font = Font(name="Calibri", size=10, bold=True, color=TEXT)

    # Freeze panes below header
    ws.freeze_panes = "A4"

    # Auto-filter on header row
    ws.auto_filter.ref = f"A3:N{3 + len(sample_data)}"


def build_services_ref_sheet(wb):
    ws = wb.create_sheet("Services Reference")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "WAYMARK HR GROUP LLC — SERVICES REFERENCE"
    title.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    title.fill = col_fill(NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = [("Service", 30), ("Description", 50), ("Typical Deliverable", 35)]
    make_header_row(ws, 2, headers, TEAL)
    ws.row_dimensions[2].height = 18

    services = [
        ("Employee Onboarding",   "Design and deliver onboarding programs and welcome packages",  "Onboarding Welcome Package (.docx)"),
        ("HR Policy Creation",    "Draft, review, and update employee handbooks and HR policies", "HR Policy Handbook (.docx)"),
        ("Workforce Analytics",   "Analyze HR data, headcount, turnover, and compensation",       "Analytics Report (.docx / .xlsx)"),
        ("Client Reporting",      "Produce professional deliverables and presentations",           "Custom Client Report (.docx)"),
        ("Compliance Review",     "Audit HR practices against federal and state requirements",     "Compliance Summary (.docx)"),
        ("Job Description Design","Create or update role-specific job descriptions",               "Job Description Set (.docx)"),
    ]

    for i, (svc, desc, deliverable) in enumerate(services, start=3):
        ws.row_dimensions[i].height = 18
        bg = ACCENT if i % 2 == 0 else WHITE
        for col, val in enumerate([svc, desc, deliverable], start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Calibri", size=10, color=TEXT)
            cell.fill = col_fill(bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border()
            ws.column_dimensions[get_column_letter(col)].width = [30, 50, 35][col - 1]


def build_status_legend_sheet(wb):
    ws = wb.create_sheet("Status Legend")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "STATUS LEGEND"
    title.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    title.fill = col_fill(NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    make_header_row(ws, 2, [("Status", 18), ("Meaning", 45)], TEAL)
    ws.row_dimensions[2].height = 18

    statuses = [
        ("Active",    "D6F0D6", "Engagement is currently in progress"),
        ("Prospect",  "FFF3CD", "Potential client — not yet engaged"),
        ("On Hold",   "FFE0CC", "Engagement paused pending client action"),
        ("Completed", "E0E0E0", "Engagement fully delivered and closed"),
    ]

    for i, (status, color, meaning) in enumerate(statuses, start=3):
        ws.row_dimensions[i].height = 18
        s_cell = ws.cell(row=i, column=1, value=status)
        s_cell.font = Font(name="Calibri", bold=True, size=10, color=TEXT)
        s_cell.fill = col_fill(color)
        s_cell.alignment = Alignment(horizontal="center", vertical="center")
        s_cell.border = thin_border()
        ws.column_dimensions["A"].width = 18

        m_cell = ws.cell(row=i, column=2, value=meaning)
        m_cell.font = Font(name="Calibri", size=10, color=TEXT)
        m_cell.fill = col_fill(color)
        m_cell.alignment = Alignment(vertical="center")
        m_cell.border = thin_border()
        ws.column_dimensions["B"].width = 45


def main():
    wb = openpyxl.Workbook()
    build_clients_sheet(wb)
    build_services_ref_sheet(wb)
    build_status_legend_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Client tracker saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
